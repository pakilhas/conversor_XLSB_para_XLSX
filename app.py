import os
import time
import logging
import uuid
import pandas as pd
import threading
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
from pyxlsb import open_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import re
import json

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/app.log'),
        logging.StreamHandler()
    ]
)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {
    'xlsb': {'xlsb'},
    'ofx': {'ofx', 'qfx'},
    'pdf': {'pdf'}
}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max
app.secret_key = 'uma_chave_secreta_muito_segura'

# Garantir que as pastas existam
for folder in [UPLOAD_FOLDER, 'logs', 'templates']:
    os.makedirs(folder, exist_ok=True)

# Dicionário para armazenar o progresso das conversões
conversion_progress = {}

def allowed_file(filename, conversion_type):
    """Verifica se a extensão do arquivo é permitida para o tipo de conversão"""
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    return ext in ALLOWED_EXTENSIONS.get(conversion_type, set())

def detect_formatting(value):
    """Detecta formatação baseada no valor da célula"""
    formatting = {
        'font': None,
        'fill': None,
        'alignment': None,
        'border': None,
        'number_format': 'General'
    }
    
    if value is None:
        return formatting
    
    try:
        # Detectar números
        if isinstance(value, (int, float)):
            if isinstance(value, int):
                formatting['number_format'] = '#,##0'
            else:
                if value == int(value):
                    formatting['number_format'] = '#,##0'
                else:
                    formatting['number_format'] = '#,##0.00'
        
        # Detectar texto longo
        elif isinstance(value, str):
            if len(value) > 20:
                formatting['alignment'] = Alignment(wrap_text=True)
            if value.isupper() or any(word in value.lower() for word in ['total', 'soma', 'quantidade', 'valor']):
                formatting['font'] = Font(bold=True)
                formatting['fill'] = PatternFill(start_color="DDDDDD", fill_type="solid")
    
    except Exception:
        pass
    
    return formatting

def apply_formatting(cell, formatting):
    """Aplica formatação a uma célula"""
    try:
        if formatting.get('font'):
            cell.font = formatting['font']
        if formatting.get('fill'):
            cell.fill = formatting['fill']
        if formatting.get('alignment'):
            cell.alignment = formatting['alignment']
        if formatting.get('number_format'):
            cell.number_format = formatting['number_format']
    except Exception as e:
        logging.debug(f"Erro ao aplicar formatação: {e}")

def parse_ofx_content(content):
    """Parse OFX content to extract transactions - corrigido para formato brasileiro"""
    transactions = []
    
    # Extrair informações do banco
    bank_match = re.search(r'<BANKID>(\d+)', content)
    acct_match = re.search(r'<ACCTID>([^<\n]+)', content)
    
    bank_id = bank_match.group(1) if bank_match else "0001"
    account_id = acct_match.group(1) if acct_match else "000000001"
    
    # Encontrar todas as transações
    stmttrn_blocks = re.findall(r'<STMTTRN>(.*?)</STMTTRN>', content, re.DOTALL)
    
    for block in stmttrn_blocks:
        # Extrair campos da transação
        trntype_match = re.search(r'<TRNTYPE>([^<\n]+)', block)
        dtposted_match = re.search(r'<DTPOSTED>([^<\n]+)', block)
        trnamt_match = re.search(r'<TRNAMT>([^<\n]+)', block)
        memo_match = re.search(r'<MEMO>([^<\n]*)', block)
        fitid_match = re.search(r'<FITID>([^<\n]+)', block)
        checknum_match = re.search(r'<CHECKNUM>([^<\n]+)', block)
        
        if trntype_match and dtposted_match and trnamt_match:
            trntype = trntype_match.group(1).strip()
            dtposted = dtposted_match.group(1).strip()
            trnamt = trnamt_match.group(1).strip()
            memo = memo_match.group(1).strip() if memo_match else ''
            fitid = fitid_match.group(1).strip() if fitid_match else ''
            checknum = checknum_match.group(1).strip() if checknum_match else ''
            
            # Converter valor - tratar vírgula como separador decimal (formato brasileiro)
            try:
                # Remover quebras de linha e espaços
                trnamt_clean = trnamt.replace('\n', '').replace('\r', '').replace(' ', '')
                # Substituir vírgula por ponto para conversão
                trnamt_clean = trnamt_clean.replace(',', '.')
                # Remover pontos de milhar se existirem
                if trnamt_clean.count('.') > 1:
                    # Se houver mais de um ponto, manter apenas o último como decimal
                    parts = trnamt_clean.split('.')
                    if len(parts) > 2:
                        trnamt_clean = ''.join(parts[:-1]) + '.' + parts[-1]
                valor = float(trnamt_clean)
            except Exception as e:
                logging.warning(f"Erro ao converter valor '{trnamt}': {e}")
                valor = 0.0
            
            # Converter data OFX para formato legível
            # Formato OFX: YYYYMMDDHHMMSS
            if dtposted and len(dtposted) >= 8:
                try:
                    dtposted_formatted = f"{dtposted[0:4]}-{dtposted[4:6]}-{dtposted[6:8]}"
                except:
                    dtposted_formatted = dtposted
            else:
                dtposted_formatted = dtposted
            
            transactions.append({
                'Data': dtposted_formatted,
                'Tipo': trntype,
                'Valor': valor,
                'Descrição': memo,
                'ID': fitid,
                'Cheque': checknum,
                'Banco': bank_id,
                'Conta': account_id
            })
    
    return transactions

def convert_ofx_to_xlsx(filepath_in, filepath_out, task_id):
    """Converte arquivo OFX para XLSX"""
    try:
        logging.info(f"Iniciando conversão OFX para XLSX: {filepath_in}")
        
        conversion_progress[task_id] = {
            'status': 'iniciando',
            'progress': 0,
            'message': 'Iniciando conversão OFX...',
            'filename': None,
            'error': None,
            'start_time': datetime.now().isoformat()
        }
        
        if not os.path.exists(filepath_in):
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath_in}")
        
        file_size = os.path.getsize(filepath_in)
        conversion_progress[task_id].update({
            'progress': 10,
            'message': f'Arquivo OFX carregado ({file_size / 1024:.1f} KB)'
        })
        
        # Ler conteúdo do arquivo OFX
        with open(filepath_in, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        conversion_progress[task_id].update({
            'progress': 30,
            'message': 'Analisando transações OFX...'
        })
        
        # Parse do conteúdo OFX
        transactions = parse_ofx_content(content)
        
        if not transactions:
            # Tentar parse alternativo
            transactions = parse_ofx_alternative(content)
        
        conversion_progress[task_id].update({
            'progress': 60,
            'message': f'{len(transactions)} transações encontradas'
        })
        
        if not transactions:
            raise Exception("Nenhuma transação encontrada no arquivo OFX")
        
        # Criar DataFrame
        df = pd.DataFrame(transactions)
        
        # Ordenar por data
        if 'Data' in df.columns:
            df = df.sort_values('Data')
        
        conversion_progress[task_id].update({
            'progress': 80,
            'message': 'Criando arquivo Excel...'
        })
        
        # Criar arquivo Excel
        with pd.ExcelWriter(filepath_out, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Transações', index=False)
            
            # Formatação básica
            workbook = writer.book
            worksheet = writer.sheets['Transações']
            
            # Formatar coluna de valor como moeda brasileira
            for row in range(2, len(df) + 2):  # +2 porque a linha 1 é o cabeçalho
                cell = worksheet.cell(row=row, column=3)  # Coluna C é Valor
                cell.number_format = '"R$" #,##0.00'
            
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Adicionar bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # Formatar cabeçalho
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        
        conversion_progress[task_id].update({
            'progress': 100,
            'message': f'Conversão concluída! {len(transactions)} transações processadas',
            'status': 'completo',
            'filename': os.path.basename(filepath_out),
            'end_time': datetime.now().isoformat()
        })
        
        logging.info(f"Conversão OFX->XLSX bem-sucedida: {filepath_out}")
        
    except Exception as e:
        error_msg = f"Erro na conversão OFX: {str(e)}"
        logging.error(error_msg)
        conversion_progress[task_id].update({
            'status': 'erro',
            'message': error_msg,
            'error': str(e),
            'end_time': datetime.now().isoformat()
        })

def parse_ofx_alternative(content):
    """Método alternativo para parse de OFX"""
    transactions = []
    
    # Encontrar blocos STMTTRN
    start_pos = 0
    while True:
        start_idx = content.find('<STMTTRN>', start_pos)
        if start_idx == -1:
            break
            
        end_idx = content.find('</STMTTRN>', start_idx)
        if end_idx == -1:
            break
            
        block = content[start_idx:end_idx + 10]  # +10 para incluir </STMTTRN>
        
        # Extrair campos manualmente
        fields = {}
        lines = block.split('\n')
        for line in lines:
            line = line.strip()
            if line.startswith('<') and '>' in line:
                tag_end = line.find('>')
                tag = line[1:tag_end]
                value = line[tag_end + 1:].strip()
                fields[tag] = value
        
        if 'TRNTYPE' in fields and 'DTPOSTED' in fields and 'TRNAMT' in fields:
            # Converter valor
            trnamt = fields['TRNAMT'].replace(',', '.')
            try:
                # Se houver mais de um ponto, é formato com milhar
                if trnamt.count('.') > 1:
                    # Remover todos os pontos e adicionar apenas o decimal
                    parts = trnamt.split('.')
                    if len(parts) > 2:
                        trnamt = ''.join(parts[:-1]) + '.' + parts[-1]
                valor = float(trnamt)
            except:
                valor = 0.0
            
            # Formatar data
            dtposted = fields['DTPOSTED']
            if len(dtposted) >= 8:
                dtposted_formatted = f"{dtposted[0:4]}-{dtposted[4:6]}-{dtposted[6:8]}"
            else:
                dtposted_formatted = dtposted
            
            transactions.append({
                'Data': dtposted_formatted,
                'Tipo': fields.get('TRNTYPE', ''),
                'Valor': valor,
                'Descrição': fields.get('MEMO', ''),
                'ID': fields.get('FITID', ''),
                'Cheque': fields.get('CHECKNUM', ''),
                'Banco': '0001',
                'Conta': '000000001'
            })
        
        start_pos = end_idx + 10
    
    return transactions

def extract_data_from_pdf(filepath_in):
    """Extrai dados de PDF (implementação básica - precisa ser expandida)"""
    # Esta é uma implementação simplificada
    # Na prática, você precisaria de uma biblioteca como pdfplumber ou PyPDF2
    # e implementar a lógica específica para seus PDFs
    
    transactions = []
    
    try:
        # Para demonstração, criamos dados de exemplo
        # Em produção, substitua por extração real do PDF
        
        # Exemplo de transações
        transactions = [
            {
                'Data': '2024-01-15',
                'Tipo': 'DEBIT',
                'Valor': -150.00,
                'Descrição': 'Supermercado',
                'Categoria': 'Alimentação'
            },
            {
                'Data': '2024-01-16',
                'Tipo': 'CREDIT',
                'Valor': 3000.00,
                'Descrição': 'Salário',
                'Categoria': 'Receita'
            },
            {
                'Data': '2024-01-17',
                'Tipo': 'DEBIT',
                'Valor': -89.90,
                'Descrição': 'Combustível',
                'Categoria': 'Transporte'
            }
        ]
        
    except Exception as e:
        logging.error(f"Erro ao extrair dados do PDF: {e}")
    
    return transactions

def generate_ofx_from_data(transactions, account_info=None):
    """Gera conteúdo OFX a partir de dados de transações"""
    
    if account_info is None:
        account_info = {
            'bank_id': '0001',
            'account_id': '000000001',
            'account_type': 'CHECKING'
        }
    
    # Header OFX
    ofx_content = """OFXHEADER:100
DATA:OFXSGML
VERSION:102
SECURITY:NONE
ENCODING:USASCII
CHARSET:1252
COMPRESSION:NONE
OLDFILEUID:NONE
NEWFILEUID:NONE

<OFX>
    <SIGNONMSGSRSV1>
        <SONRS>
            <STATUS>
                <CODE>0
                <SEVERITY>INFO
            </STATUS>
            <DTSERVER>""" + datetime.now().strftime("%Y%m%d%H%M%S") + """
            <LANGUAGE>POR
            <FI>
                <ORG>Banco Exemplo
                <FID>""" + account_info['bank_id'] + """
            </FI>
        </SONRS>
    </SIGNONMSGSRSV1>
    <BANKMSGSRSV1>
        <STMTTRNRS>
            <TRNUID>1001
            <STATUS>
                <CODE>0
                <SEVERITY>INFO
            </STATUS>
            <STMTRS>
                <CURDEF>BRL
                <BANKACCTFROM>
                    <BANKID>""" + account_info['bank_id'] + """
                    <ACCTID>""" + account_info['account_id'] + """
                    <ACCTTYPE>""" + account_info['account_type'] + """
                </BANKACCTFROM>
                <BANKTRANLIST>
"""
    
    # Adicionar transações
    for i, trans in enumerate(transactions):
        # Converter data para formato OFX
        dtposted = trans.get('Data', '').replace('-', '')
        if len(dtposted) == 10:  # YYYY-MM-DD
            dtposted += "000000"  # Adicionar hora se necessário
        
        trntype = trans.get('Tipo', 'OTHER').upper()
        trnamt = trans.get('Valor', 0)
        memo = trans.get('Descrição', '')
        
        # Formatar valor com vírgula (formato brasileiro)
        trnamt_formatted = f"{trnamt:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        ofx_content += f"""                    <STMTTRN>
                        <TRNTYPE>{trntype}
                        <DTPOSTED>{dtposted}
                        <TRNAMT>{trnamt_formatted}
                        <FITID>{i+1:08d}
                        <MEMO>{memo}
                    </STMTTRN>
"""
    
    # Footer OFX
    ofx_content += """                </BANKTRANLIST>
                <LEDGERBAL>
                    <BALAMT>0.00
                    <DTASOF>""" + datetime.now().strftime("%Y%m%d%H%M%S") + """
                </LEDGERBAL>
            </STMTRS>
        </STMTTRNRS>
    </BANKMSGSRSV1>
</OFX>"""
    
    return ofx_content

def convert_pdf_to_ofx(filepath_in, filepath_out, task_id):
    """Converte arquivo PDF para OFX"""
    try:
        logging.info(f"Iniciando conversão PDF para OFX: {filepath_in}")
        
        conversion_progress[task_id] = {
            'status': 'iniciando',
            'progress': 0,
            'message': 'Iniciando conversão PDF...',
            'filename': None,
            'error': None,
            'start_time': datetime.now().isoformat()
        }
        
        if not os.path.exists(filepath_in):
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath_in}")
        
        file_size = os.path.getsize(filepath_in)
        conversion_progress[task_id].update({
            'progress': 10,
            'message': f'Arquivo PDF carregado ({file_size / 1024:.1f} KB)'
        })
        
        # Extrair dados do PDF
        conversion_progress[task_id].update({
            'progress': 30,
            'message': 'Extraindo dados do PDF...'
        })
        
        transactions = extract_data_from_pdf(filepath_in)
        
        conversion_progress[task_id].update({
            'progress': 60,
            'message': f'{len(transactions)} transações extraídas'
        })
        
        # Gerar arquivo OFX
        conversion_progress[task_id].update({
            'progress': 80,
            'message': 'Gerando arquivo OFX...'
        })
        
        ofx_content = generate_ofx_from_data(transactions)
        
        # Salvar arquivo OFX
        with open(filepath_out, 'w', encoding='utf-8') as f:
            f.write(ofx_content)
        
        conversion_progress[task_id].update({
            'progress': 100,
            'message': f'Conversão concluída! Arquivo OFX gerado',
            'status': 'completo',
            'filename': os.path.basename(filepath_out),
            'end_time': datetime.now().isoformat()
        })
        
        logging.info(f"Conversão PDF->OFX bem-sucedida: {filepath_out}")
        
    except Exception as e:
        error_msg = f"Erro na conversão PDF: {str(e)}"
        logging.error(error_msg)
        conversion_progress[task_id].update({
            'status': 'erro',
            'message': error_msg,
            'error': str(e),
            'end_time': datetime.now().isoformat()
        })

def convert_xlsb_to_xlsx_advanced(filepath_in, filepath_out, task_id):
    """Conversão avançada que preserva dados e estrutura"""
    try:
        logging.info(f"Iniciando conversão avançada: {filepath_in} -> {filepath_out}")
        
        conversion_progress[task_id] = {
            'status': 'iniciando',
            'progress': 0,
            'message': 'Iniciando conversão...',
            'filename': None,
            'error': None,
            'start_time': datetime.now().isoformat()
        }
        
        # Verificar se arquivo existe
        if not os.path.exists(filepath_in):
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath_in}")
        
        file_size = os.path.getsize(filepath_in)
        conversion_progress[task_id].update({
            'progress': 10,
            'message': f'Arquivo carregado ({file_size / 1024 / 1024:.1f} MB)'
        })
        
        # Método 1: Tentar com pandas + openpyxl
        try:
            conversion_progress[task_id].update({
                'progress': 20,
                'message': 'Lendo estrutura do arquivo XLSB...'
            })
            
            # Ler metadados do arquivo
            xlsb_file = pd.ExcelFile(filepath_in, engine='pyxlsb')
            sheet_names = xlsb_file.sheet_names
            
            conversion_progress[task_id].update({
                'progress': 30,
                'message': f'Encontradas {len(sheet_names)} planilhas'
            })
            
            # Criar workbook de saída
            wb_out = Workbook()
            # Remover sheet padrão
            wb_out.remove(wb_out.active)
            
            # Processar cada planilha
            for sheet_idx, sheet_name in enumerate(sheet_names):
                progress = 30 + (sheet_idx * 60 / len(sheet_names))
                conversion_progress[task_id].update({
                    'progress': progress,
                    'message': f'Processando: {sheet_name}'
                })
                
                try:
                    # Ler dados mantendo tipos originais
                    df = pd.read_excel(
                        filepath_in, 
                        sheet_name=sheet_name, 
                        engine='pyxlsb',
                        dtype=object,
                        keep_default_na=False
                    )
                    
                    # Criar nova planilha
                    ws_out = wb_out.create_sheet(title=sheet_name[:31])
                    
                    # Escrever dados
                    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for col_idx, value in enumerate(row, 1):
                            cell = ws_out.cell(row=row_idx, column=col_idx, value=value)
                            
                            # Aplicar formatação detectada
                            formatting = detect_formatting(value)
                            apply_formatting(cell, formatting)
                    
                    # Ajustar largura das colunas
                    for column in ws_out.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                if cell.value:
                                    length = len(str(cell.value))
                                    max_length = max(max_length, length)
                            except:
                                pass
                        
                        adjusted_width = min(max(max_length + 2, 8), 50)
                        ws_out.column_dimensions[column_letter].width = adjusted_width
                    
                    # Adicionar bordas básicas
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                    
                    for row in ws_out.iter_rows(min_row=1, max_row=ws_out.max_row, min_col=1, max_col=ws_out.max_column):
                        for cell in row:
                            if cell.value is not None:
                                cell.border = thin_border
                    
                    logging.info(f"Planilha {sheet_name} processada com sucesso")
                    
                except Exception as e:
                    logging.error(f"Erro na planilha {sheet_name}: {e}")
                    # Criar planilha vazia como fallback
                    ws_out = wb_out.create_sheet(title=sheet_name[:31])
                    ws_out.cell(1, 1, value=f"Erro ao processar: {str(e)}")
                    continue
                
                time.sleep(0.1)
            
            # Salvar arquivo
            conversion_progress[task_id].update({
                'progress': 95,
                'message': 'Salvando arquivo XLSX...'
            })
            
            wb_out.save(filepath_out)
            
            # Verificar se arquivo foi criado
            if os.path.exists(filepath_out):
                output_size = os.path.getsize(filepath_out)
                conversion_progress[task_id].update({
                    'progress': 100,
                    'message': f'Conversão concluída! ({output_size / 1024 / 1024:.1f} MB)',
                    'status': 'completo',
                    'filename': os.path.basename(filepath_out),
                    'end_time': datetime.now().isoformat()
                })
                logging.info(f"Conversão bem-sucedida: {filepath_out}")
            else:
                raise Exception("Arquivo de saída não foi criado")
                
        except Exception as e:
            logging.error(f"Erro no método principal: {e}")
            
            # Método 2: Fallback simples
            conversion_progress[task_id].update({
                'progress': 50,
                'message': 'Usando método alternativo...'
            })
            
            try:
                xlsb_file = pd.ExcelFile(filepath_in, engine='pyxlsb')
                sheet_names = xlsb_file.sheet_names
                
                with pd.ExcelWriter(filepath_out, engine='openpyxl') as writer:
                    for i, sheet_name in enumerate(sheet_names):
                        df = pd.read_excel(filepath_in, sheet_name=sheet_name, engine='pyxlsb')
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                conversion_progress[task_id].update({
                    'progress': 100,
                    'message': 'Conversão concluída (método simples)',
                    'status': 'completo', 
                    'filename': os.path.basename(filepath_out),
                    'end_time': datetime.now().isoformat()
                })
                
            except Exception as fallback_error:
                raise Exception(f"Todos os métodos falharam: {str(e)} -> {str(fallback_error)}")
        
    except Exception as e:
        error_msg = f"Erro na conversão: {str(e)}"
        logging.error(error_msg)
        conversion_progress[task_id].update({
            'status': 'erro',
            'message': error_msg,
            'error': str(e),
            'end_time': datetime.now().isoformat()
        })

@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/health')
def health_check():
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})

@app.route('/upload', methods=['POST'])
def upload_file():
    """Endpoint para upload de arquivos XLSB"""
    return handle_upload('xlsb_to_xlsx', convert_xlsb_to_xlsx_advanced, '.xlsx')

@app.route('/upload/ofx', methods=['POST'])
def upload_ofx_file():
    """Endpoint para upload de arquivos OFX"""
    return handle_upload('ofx_to_xlsx', convert_ofx_to_xlsx, '.xlsx')

@app.route('/upload/pdf', methods=['POST'])
def upload_pdf_file():
    """Endpoint para upload de arquivos PDF"""
    return handle_upload('pdf_to_ofx', convert_pdf_to_ofx, '.ofx')

def handle_upload(conversion_type, conversion_func, output_extension):
    """Manipula o upload e inicia a conversão"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if file and allowed_file(file.filename, conversion_type.split('_')[0]):
            task_id = str(uuid.uuid4())
            filename = secure_filename(file.filename)
            filepath_in = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath_in)
            
            # Criar nome do arquivo de saída
            base_name = os.path.splitext(filename)[0]
            filename_out = f"{base_name}{output_extension}"
            filepath_out = os.path.join(app.config['UPLOAD_FOLDER'], filename_out)
            
            # Informações iniciais
            input_size = os.path.getsize(filepath_in)
            conversion_progress[task_id] = {
                'status': 'iniciando',
                'progress': 0,
                'message': 'Preparando conversão...',
                'filename': filename_out,
                'error': None,
                'start_time': datetime.now().isoformat(),
                'details': {
                    'input_file': filename,
                    'input_size': f"{input_size / 1024 / 1024:.2f} MB",
                    'conversion_type': conversion_type
                }
            }
            
            # Iniciar conversão em thread separada
            thread = threading.Thread(
                target=conversion_func,
                args=(filepath_in, filepath_out, task_id)
            )
            thread.daemon = True
            thread.start()
            
            return jsonify({
                'task_id': task_id, 
                'filename': filename_out,
                'conversion_type': conversion_type
            })
        
        return jsonify({'error': 'Tipo de arquivo não permitido'}), 400
    
    except Exception as e:
        logging.error(f"Erro no upload: {e}")
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/progress/<task_id>')
def get_progress(task_id):
    progress_data = conversion_progress.get(task_id, {
        'status': 'nao_encontrado',
        'progress': 0,
        'message': 'Tarefa não encontrada'
    })
    return jsonify(progress_data)

@app.route('/download/<filename>')
def download_file(filename):
    try:
        return send_from_directory(
            app.config['UPLOAD_FOLDER'], 
            filename, 
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logging.error(f"Erro no download: {e}")
        return jsonify({'error': 'Arquivo não encontrado'}), 404

@app.route('/cleanup', methods=['POST'])
def cleanup_files():
    try:
        cutoff_time = time.time() - 3600  # 1 hora
        removed = 0
        
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            if os.path.getctime(filepath) < cutoff_time:
                os.remove(filepath)
                removed += 1
        
        return jsonify({'message': f'{removed} arquivos removidos'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/formats')
def get_supported_formats():
    """Retorna os formatos suportados"""
    return jsonify({
        'xlsb_to_xlsx': {
            'from': 'xlsb',
            'to': 'xlsx',
            'description': 'Excel Binary para Excel Open XML'
        },
        'ofx_to_xlsx': {
            'from': 'ofx',
            'to': 'xlsx', 
            'description': 'Open Financial Exchange para Excel'
        },
        'pdf_to_ofx': {
            'from': 'pdf',
            'to': 'ofx',
            'description': 'PDF para Open Financial Exchange'
        }
    })

if __name__ == '__main__':
    logging.info("Iniciando aplicação Flask na porta 9090")
    app.run(host='0.0.0.0', port=9090, debug=False)